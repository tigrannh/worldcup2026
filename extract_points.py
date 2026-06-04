import openpyxl

def get_points_info(file_path):
    print(f"\n--- Extracting Points from: {file_path} ---")
    wb = openpyxl.load_workbook(file_path, data_only=True) # Get values, not formulas for points
    
    if "Game Setup" in wb.sheetnames:
        ws = wb["Game Setup"]
        print("Game Setup Point Values:")
        # Look around where we saw 'GROUP STAGES Point System' [B5]
        for r in range(5, 25):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
            if any(row_vals):
                print(f"Row {r}: {row_vals}")

    if "Players" in wb.sheetnames:
        ws = wb["Players"]
        print("\nPlayers Sheet Sample:")
        for r in range(1, 10):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
            print(f"Row {r}: {row_vals}")

get_points_info("Euro 2024 - Phoenixes.xlsx")
get_points_info("Euro 2024_Players.xlsm")
