import openpyxl
import json

def exhaustive_scan(file_path):
    print(f"DEBUG: Starting exhaustive scan of {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    report = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        report[sheet_name] = {
            "dimensions": f"{sheet.min_row}:{sheet.max_row} x {sheet.min_column}:{sheet.max_column}",
            "key_formulas": [],
            "headers": []
        }
        
        # Scan headers (first 2 rows)
        for r in range(1, 3):
            headers = [str(sheet.cell(row=r, column=c).value) for c in range(1, min(sheet.max_column, 20) + 1) if sheet.cell(row=r, column=c).value]
            if headers:
                report[sheet_name]["headers"].append(headers)
        
        # Scan for interesting formulas throughout the sheet
        # We limit the search to avoid huge outputs, but look for specific keywords
        count = 0
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 200), min_col=1, max_col=min(sheet.max_column, 30)):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                if val.startswith("="):
                    # Only record unique or complex formulas
                    if any(key in val.upper() for key in ["IF", "SUM", "VLOOKUP", "MATCH", "INDEX", "RANK", "COUNTIFS"]):
                        if count < 50: # Cap per sheet for summary
                            report[sheet_name]["key_formulas"].append(f"[{cell.coordinate}] {val}")
                            count += 1

    # Specifically look at "Matches" and "Player Game Board" which are the "Engine"
    engine_report = ""
    for sheet_name in ["Matches", "Player Game Board", "Player Scoreboard", "Player Leaderboard"]:
        if sheet_name in wb.sheetnames:
            engine_report += f"\n--- {sheet_name} ENGINE LOGIC ---\n"
            sheet = wb[sheet_name]
            # Sample formulas from middle rows where calculation happens
            for r in range(8, 15):
                row_data = [f"[{sheet.cell(row=r, column=c).coordinate}] {sheet.cell(row=r, column=c).value}" for c in range(10, 25)]
                engine_report += " | ".join(row_data) + "\n"

    return report, engine_report

report, engine = exhaustive_scan("Euro 2024 - Phoenixes.xlsx")

print("\n--- STRUCTURE REPORT ---")
for s, data in report.items():
    print(f"\nSheet: {s} ({data['dimensions']})")
    print(f"Headers: {data['headers']}")
    print(f"Formula Samples: {data['key_formulas'][:5]}")

print(engine)
