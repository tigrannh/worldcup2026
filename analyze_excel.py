import pandas as pd
import openpyxl

def analyze_excel(file_path):
    print(f"\n--- Analyzing: {file_path} ---")
    try:
        # Load workbook to see sheet names and potentially formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)
        print(f"Sheets: {wb.sheetnames}")
        
        # Check first few sheets for formula-like patterns in common areas
        for sheet_name in wb.sheetnames[:3]:
            sheet = wb[sheet_name]
            print(f"\nScanning Sheet: {sheet_name}")
            
            # Look for headers or formulas that calculate points
            # We'll sample some rows/cols
            for row in sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=15):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # If it looks like a formula or a point-related header
                        if "=" in str(cell.value) or any(x in str(cell.value).lower() for x in ["point", "score", "predict", "logic"]):
                            print(f"  [{cell.coordinate}] {cell.value}")
                            
    except Exception as e:
        print(f"Error: {e}")

files = ["Euro 2024 - Phoenixes.xlsx", "Euro 2024_Players.xlsm"]
for f in files:
    analyze_excel(f)
